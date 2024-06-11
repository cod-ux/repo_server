import os 
import xlwings as xl
import streamlit as st
import pandas as pd
from openpyxl import load_workbook

def re_upload():
    st.session_state.uploaded_file = None
    st.session_state.file_path = None
    st.cache_data.clear()

def copy_excel_locally(file):
    fname, ext = os.path.splitext(file.name)
    new_fname = f"{fname}_copy{ext}"
    old_fname = f"{fname}{ext}"
    file_root = os.path.join("/Users/suryaganesan/vscode/ml/projects/reporter/uploads", new_fname)
    original_path = os.path.join("/Users/suryaganesan/vscode/ml/projects/reporter/", old_fname)
    with open(file_root, "wb") as local_file:
        local_file.write(file.read())

    file_path = f"/Users/suryaganesan/vscode/ml/projects/reporter/uploads/{new_fname}"

    return file_path

def copy_excel_locally_from_path(file):
    fname, ext = os.path.splitext(file)
    new_fname = f"{fname}_copy{ext}"
    file_root = os.path.join("/Users/suryaganesan/vscode/ml/projects/reporter/uploads", new_fname)
    with open(file_root, "wb") as local_file:
        local_file.write(file.read())
    st.success(f"File saved as {new_fname}")

    file_path = f"/Users/suryaganesan/vscode/ml/projects/reporter/uploads/{new_fname}"

    return file_path


def handle_duplicate_columns(columns):
    counts = {}
    new_columns = []

    for col_name in columns:
        if col_name in counts:
            counts[col_name] += 1
            new_columns.append(f"{col_name}_{counts[col_name]}")
        else:
            counts[col_name] = 0
            new_columns.append(col_name)
    
    return new_columns


@st.cache_data()
def load_sheets_to_dfs(file_path):
    app = xl.App(visible=False)
    wb = app.books.open(file_path)
    dfs = []
    for sheet in wb.sheets:
        df = sheet.used_range.options(pd.DataFrame, header=True, index=False).value
        df.columns = handle_duplicate_columns(df.columns)
        dfs.append(df)

    sheet_names = [sheet.name for sheet in wb.sheets]
    wb.save()
    wb.close()
    app.quit()
    return dfs, sheet_names


def save_sheets(path):
    app = xl.App(visible=False)
    book = app.books.open(path)
    book.save()
    book.close()
    app.kill()
    return True

def unmerge_sheets(file_path):
    book = load_workbook(file_path)
    for sheet in book.sheetnames:
        ws = book[sheet]
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range in list(ws.merged_cells.ranges):
                ws.unmerge_cells(str(merged_range))
                print("Merged Range: ", merged_range)

    book.save(file_path)
    book.close()

def get_binary(file_path):
    with open(file_path, "rb") as f:
        bn = f.read()
        
    return bn

def undo(file_path):
    try:
        last_state = st.session_state.state_stack.pop()
        if last_state:
            with open(file_path, "wb") as f:
                f.write(last_state)

            st.cache_data.clear()

    except Exception as e:
        print("Error with undo action: ", e)
        return
